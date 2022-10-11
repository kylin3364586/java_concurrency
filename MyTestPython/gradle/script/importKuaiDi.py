#导入快递数据

import xlrd
import openpyxl
import pymysql
#打开数据所在的工作簿，以及选择存有数据的工作表
book = openpyxl.load_workbook("C:\\Users\\admin\\Desktop\\快递100快递公司标准编码-20210716104728.xlsx")
sheet = book.get_sheet_by_name("快递100快递公司编码表")

#建立一个MySQL连接
conn = pymysql.connect(
    host='192.168.20.47',
    user='root',
    passwd='Rchz2020@',
    db='test-yijia-app',
    port=3306,
    charset='utf8'
)
# 获得游标
cur = conn.cursor()
# 创建插入SQL语句
query = 'insert into yj_logistics (shipping_name,tracking_number) values (%s, %s)'
# 创建一个for循环迭代读取xls文件每行数据的, 从第二行开始是要跳过标题行
for r in range(1, sheet.max_row +1):
    shipping_name = sheet.cell(r,1).value
    tracking_number = sheet.cell(r,2).value
    values = (shipping_name,tracking_number)
    print(values)
    # 执行sql语句
    cur.execute(query, values)
cur.close()
conn.commit()
conn.close()
columns = str(sheet.ncols)
rows = str(sheet.nrows)
print ("导入 " +columns + " 列 " + rows + " 行数据到MySQL数据库!")